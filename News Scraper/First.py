from datetime import date
import os
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

url = 'https://www.bbc.com/'
today_date = date.today().strftime('%Y-%m-%d')  # Format date as a string for consistent sheet naming
print(today_date)
news_data = []

def scrape_articles():
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    articles = soup.find_all('h2')
    for article in articles:
        headline = article.text.strip()
        news_data.append({'headline': headline})

    print("Successfully scraped articles")
    

def ensure_sheet_exists(file_path):
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = str(today_date)
        workbook.save(file_path)
        print(f'Created new Excel file and sheet at {file_path}')
    else:
        workbook = load_workbook(file_path)
        if str(today_date) not in workbook.sheetnames:
            workbook.create_sheet(title=str(today_date))
            workbook.save(file_path)
            print(f'New sheet for {today_date} created in existing workbook.')
        else:
            print(f'Sheet for {today_date} already exists.')
    return workbook

def enter_article_to_excel(file_path):
    workbook = ensure_sheet_exists(file_path)
    sheet = workbook[str(today_date)]  # Access the correct sheet based on today's date
    sheet['a1'] = 'Article Headlines'
    row = 2
    for item in news_data:
        cell = f'A{row}'  # Column A
        sheet[cell] = item['headline']
        row += 1
    workbook.save(file_path)
    print('Headlines saved to Excel.')

# Example usage:
scrape_articles()
import os
cwd = os.getcwd()
print(cwd)
enter_article_to_excel(f'{cwd}\BBC_News_Headlines.xlsx')
